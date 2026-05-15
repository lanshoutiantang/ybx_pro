import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def extract_numeric_value(filename):
    """从txt文件名中提取用于排序的数值（取第二个数值）"""
    # 处理overall_pr特殊情况
    if filename == "overall_pr.txt":
        return -1  # 让overall_pr排最前面（不受reverse影响）
    # 提取pr_xxx_xxx中的第二个数值
    parts = filename.replace("pr_", "").replace(".txt", "").split("_")
    try:
        # 取第二个数值作为排序依据（如pr_200_850取850，pr_850_10000取10000）
        return int(parts[1])
    except (IndexError, ValueError):
        return 0

def generate_pr_xlsx(versions, base_parent_path, output_file):
    """
    多版本+多二级目录数据汇总到**单个Excel文件**，**不同二级目录对应不同Sheet**
    :param versions: 版本列表，如 ["s018_1920", "s019_1920"]
    :param base_parent_path: 所有版本的父目录
    :param output_file: 最终输出的xlsx文件路径，如 "./s019_weak-mining-qat/ks_results_所有目录汇总.xlsx"
    """
    # 要处理的二级目录（Sheet名对应此名称）
    level2_dirs = [
        "carry_head", "carry_no_mask_head", "sps_head", "sps_no_mask_head", "carry_waist",
        "carry_no_mask_waist","navigate_back", "navigate_head", "navigate_waist",
        "carry","carry_shiyan","navigate","sps"
    ]
    # 固定要处理的txt文件
    txt_files = ['pr_200_850.txt', 'pr_850_10000.txt', 'pr_200_1000.txt',
                 'pr_1000_2000.txt', 'pr_2000_5000.txt', 'overall_pr.txt']

    # 初始化：按二级目录存储所有版本的数据（原有逻辑不变）
    dir_all_data = defaultdict(lambda: {
        "results": {},  # 存储 {版本_ks_* 文件名: "precision recall"}
        "existing_files": set()  # 存储该二级目录下所有版本中存在的文件
    })

    # 第一步：遍历所有版本，收集所有数据（原有逻辑完全不变）
    for version in versions:
        print(f"===== 处理版本: {version} =====")
        version_base_path = os.path.join(base_parent_path, version)

        for level2_dir in level2_dirs:
            src_base_path = os.path.join(version_base_path, level2_dir)
            if not os.path.isdir(src_base_path):
                print(f"[{version}] 源目录不存在: {src_base_path}")
                continue

            print(f"[{version}] 处理二级目录: {level2_dir}")
            # 遍历每个ks目录（0-5）
            for ks in range(6):
                ks_src_path = os.path.join(src_base_path, f"ks_{ks}")
                if not os.path.isdir(ks_src_path):
                    continue

                # 遍历每个txt文件
                for txt_file in txt_files:
                    src_path = os.path.join(ks_src_path, txt_file)
                    if os.path.exists(src_path):
                        dir_all_data[level2_dir]["existing_files"].add(txt_file)
                    if not os.path.exists(src_path):
                        continue

                    try:
                        # 读取文件提取AVERAGE行的Precision和Recall
                        with open(src_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                line = line.rstrip('\n')
                                if line and "AVERAGE" in line:
                                    parts = line.strip().split()
                                    precision = None
                                    recall = None
                                    # 提取Precision和Recall值
                                    for part in parts:
                                        if part.startswith("Precision="):
                                            precision = part.split("=")[1]
                                        elif part.startswith("Recall="):
                                            recall = part.split("=")[1]
                                    # 保存结果
                                    if precision and recall:
                                        name = txt_file.replace(".txt", "")
                                        key = f"{version}_ks_{ks} {name}"
                                        dir_all_data[level2_dir]["results"][key] = f"{precision} {recall}"
                    except Exception as e:
                        print(f"[{version}] 处理 {src_path} 时出错: {e}")
                        continue

    # 第二步：创建单个Excel工作簿，每个二级目录对应一个Worksheet
    wb = Workbook()
    # 删除openpyxl默认创建的Sheet1
    wb.remove(wb.active)
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # 遍历每个二级目录，创建对应的Sheet并写入数据
    for level2_dir in level2_dirs:
        dir_data = dir_all_data[level2_dir]
        results = dir_data["results"]
        existing_files = dir_data["existing_files"]

        if not results:
            print(f"二级目录 {level2_dir} 无有效数据，跳过创建Sheet")
            continue

        # 处理Sheet名：Excel Sheet名最大31字符，移除非法字符（/:*?[]\）
        sheet_name = level2_dir
        illegal_chars = [':', '/', '*', '?', '[', ']', '\\']
        for c in illegal_chars:
            sheet_name = sheet_name.replace(c, '_')
        sheet_name = sheet_name[:31]  # 截断过长名称
        # 创建新Sheet
        ws = wb.create_sheet(title=sheet_name)
        print(f"为二级目录 {level2_dir} 创建Sheet: {sheet_name}")

        # 构建列名（原有逻辑不变）
        columns_set = ["version", "ks"]
        sorted_file_names = sorted(existing_files, key=lambda x: extract_numeric_value(x), reverse=False)
        for fn in sorted_file_names:
            pre_col = fn.replace(".txt", "") + "_precision"
            rec_col = fn.replace(".txt", "") + "_recall"
            if pre_col not in columns_set:
                columns_set.append(pre_col)
            if rec_col not in columns_set:
                columns_set.append(rec_col)

        # 解析结果到分组数据（原有逻辑不变）
        version_ks_groups = defaultdict(lambda: defaultdict(dict))
        for item in results:
            version_ks_part, name = item.split(" ", 1)
            version_part, ks_part = version_ks_part.split("_ks_", 1)
            parts = results[item].split()
            precision = parts[0]
            recall = parts[1]
            version_ks_groups[version_part][ks_part][name + "_precision"] = precision
            version_ks_groups[version_part][ks_part][name + "_recall"] = recall

        # 第一步：写入列名（表头）
        for col_idx, col_name in enumerate(columns_set, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 第二步：写入数据行
        row_idx = 2  # 从第二行开始写入数据
        for version in sorted(versions):
            ks_dict = version_ks_groups.get(version, {})
            sorted_ks = sorted(ks_dict.keys(), key=lambda x: int(x))
            for ks in sorted_ks:
                # 写入version和ks列
                ws.cell(row=row_idx, column=1, value=version)
                ws.cell(row=row_idx, column=2, value=f"ks_{ks}")
                # 写入其他列数据
                for col_idx, col_name in enumerate(columns_set, 1):
                    if col_name not in ["version", "ks"]:
                        cell_value = ks_dict[ks].get(col_name, "")
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                row_idx += 1

        # 可选优化：自动调整列宽（根据内容长度）
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50，避免过宽
            ws.column_dimensions[column_letter].width = adjusted_width

        print(f"  Sheet {sheet_name} 写入完成，共 {row_idx-2} 行数据，{len(columns_set)} 列")

    # 保存最终的Excel文件（单文件多Sheet）
    wb.save(output_file)
    print(f"\n✅ 所有数据汇总完成！单一Excel文件已生成：")
    print(f"   路径：{os.path.abspath(output_file)}")
    print(f"   包含Sheet数：{len(wb.sheetnames)}")
    print(f"   包含Sheet：{wb.sheetnames}")

# 配置参数
if __name__ == "__main__":
    # 要汇总的版本列表
    versions = [
        # "1280/s0110-weak-mining-qat",
        # "1920/s0110-weak-mining-qat"
        # "s0.1.9_mining_gama1.05_1280",
        # "s0.1.9_mining_gama1.05_1920",
        # "s0.1.9_mono_1280",
        # "s0.1.9_mono_1920",
        # "s019_1280",
        # "s019_1920",
        # "s019_weak-mining-qat_1280",
        # "s019_weak-mining-qat_1920",
        # "s0.1.11_mono_navigate_hecheng",
        # "s0.1.11_mono",
        # "s0111_1280",
        # "s0111_1920",
        # "s0.1.14_origin_1280",
        # "s0.1.14_origin_1920",
        # "s0.1.14_dla_1280",
        # "s0.1.14_dla_1920",
        "s0.1.14_1280",
        "s0.1.14_1920",
    ]
    # 所有版本的父目录（去掉版本部分）
    base_parent_path = "D:\Program Files\JetBrains\pythonProject\stereo\pr_test1"
    # 单一Excel输出文件路径（核心修改：从目录改为具体xlsx文件）
    output_file = "./ks_results_所有目录汇总.xlsx"

    # 执行生成逻辑
    generate_pr_xlsx(versions, base_parent_path, output_file)
