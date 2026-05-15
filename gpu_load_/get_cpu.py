import os
from datetime import datetime, timedelta
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 全局定义核心列（解决作用域问题）
CORE_COLS = ["时间", "瞬时%CPU", "瞬时%MEM"]


def parse_time_str(time_str: str) -> datetime:
    """解析时间字符串为datetime对象，兼容格式：YYYY-MM-DD HH:MM:SS"""
    try:
        return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        # 兼容常见的时间格式错误（如多余空格）
        time_str = time_str.strip().replace("  ", " ")
        return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")


def calculate_time_offset(original_base: str, target_base: str) -> timedelta:
    """计算原始基准时间与目标基准时间的差值（用于所有时间偏移）"""
    try:
        original_dt = parse_time_str(original_base)
        target_dt = parse_time_str(target_base)
        return target_dt - original_dt
    except Exception as e:
        raise ValueError(f"时间基准计算失败：{e}")


def normalize_header(header: str) -> str:
    """标准化表头（去除空格/特殊符号，用于模糊匹配）"""
    return header.strip() \
        .replace(" ", "") \
        .replace("%", "百分比") \
        .replace("（", "") \
        .replace("）", "")


def process_single_txt(txt_path: str, time_offset: timedelta) -> List[Dict]:
    """处理单个TXT文件，按时间偏移修正所有时间字段（增强兼容性）"""
    processed_data = []
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f if line.strip()]  # 过滤空行
    except UnicodeDecodeError:
        # 兼容GBK编码的TXT文件
        with open(txt_path, 'r', encoding='gbk') as f:
            lines = [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"❌ 读取文件失败 {txt_path}：{e}")
        return processed_data

    if not lines:
        print(f"⚠️ 文件为空 {txt_path}，跳过")
        return processed_data

    # 解析表头并标准化
    raw_header = lines[0].split(',')
    norm_header = [normalize_header(col) for col in raw_header]
    norm_core_cols = [normalize_header(col) for col in CORE_COLS]

    # 宽松校验：匹配核心列（允许表头格式差异）
    header_mapping = {}  # 标准列名 → 原始列索引
    missing_cols = []
    for core_col, norm_core in zip(CORE_COLS, norm_core_cols):
        if norm_core in norm_header:
            header_mapping[core_col] = norm_header.index(norm_core)
        else:
            missing_cols.append(core_col)

    if missing_cols:
        print(f"⚠️ 文件 {txt_path} 缺失核心列 {missing_cols}，跳过")
        return processed_data

    # 处理数据行
    for line_num, line in enumerate(lines[1:], start=2):
        try:
            # 分割数据（避免启动命令中的逗号/空格被拆分）
            row_data = line.split(',', maxsplit=len(raw_header) - 1)
            if len(row_data) < len(raw_header):
                print(f"⚠️ {txt_path} 第{line_num}行列数不足，跳过")
                continue

            # 构建数据字典
            row_dict = {}
            # 处理核心列
            for core_col in CORE_COLS:
                idx = header_mapping[core_col]
                col_value = row_data[idx].strip() if idx < len(row_data) else ""

                if core_col == "时间":
                    # 时间偏移处理
                    original_dt = parse_time_str(col_value)
                    new_dt = original_dt + time_offset
                    row_dict[core_col] = new_dt.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    row_dict[core_col] = col_value

            # 处理启动命令（如果有）
            if "启动命令" in raw_header:
                cmd_idx = raw_header.index("启动命令")
                row_dict["启动命令"] = row_data[cmd_idx].strip() if cmd_idx < len(row_data) else ""

            processed_data.append(row_dict)

        except Exception as e:
            print(f"⚠️ {txt_path} 第{line_num}行处理失败：{e}，跳过")
            continue

    print(f"✅ 处理完成 {txt_path}：提取到 {len(processed_data)} 行数据")
    return processed_data


def batch_process_to_excel(
        txt_dir: str,
        original_base_time: str,
        target_base_time: str,
        output_excel: str
):
    """批量处理TXT文件并输出到Excel（带格式美化）"""
    # 计算时间偏移量
    try:
        time_offset = calculate_time_offset(original_base_time, target_base_time)
        print(f"📌 时间偏移量：{time_offset}（所有时间将增加该时长）")
    except Exception as e:
        print(f"❌ 时间基准计算失败：{e}")
        return

    # 收集所有数据
    all_data = []
    has_command_col = False  # 标记是否有启动命令列

    # 遍历目录下所有TXT文件
    for filename in os.listdir(txt_dir):
        if not filename.lower().endswith('.log'):
            continue
        txt_path = os.path.join(txt_dir, filename)
        file_data = process_single_txt(txt_path, time_offset)
        if file_data:
            all_data.extend(file_data)
            # 检查是否有启动命令列
            if "启动命令" in file_data[0]:
                has_command_col = True

    if not all_data:
        print("❌ 未提取到任何有效数据！")
        return

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "处理后数据"

    # 定义表头
    headers = CORE_COLS.copy()
    if has_command_col:
        headers.append("启动命令")

    # 写入表头并美化格式
    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align
        # 自动调整列宽
        ws.column_dimensions[chr(64 + col_idx)].width = 20 if header == "启动命令" else 15

    # 写入数据
    for row_idx, row_dict in enumerate(all_data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header, ""))

    # 保存Excel文件
    try:
        wb.save(output_excel)
        print(f"🎉 处理完成！结果已保存至：{output_excel}")
    except PermissionError:
        print(f"❌ 保存失败：{output_excel} 已被打开，请关闭后重试")
    except Exception as e:
        print(f"❌ 保存Excel失败：{e}")



if __name__ == "__main__":
    # ===================== 配置参数（按需修改）=====================
    TXT_DIR = "/home/ubt/下载/s0.1.5-qat/xingneng/v0.1.5.0"  # TXT文件所在目录
    ORIGINAL_BASE_TIME = "1970-01-01 11:41:17"  # 原始基准时间
    TARGET_BASE_TIME = "2025-12-29 18:35:34"  # 目标基准时间
    OUTPUT_EXCEL = "v0.1.5.0cpu处理后数据_20251229.xlsx"  # 输出Excel文件名

    # 执行批量处理
    batch_process_to_excel(TXT_DIR, ORIGINAL_BASE_TIME, TARGET_BASE_TIME, OUTPUT_EXCEL)